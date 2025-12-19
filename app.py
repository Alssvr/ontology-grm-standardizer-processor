import gradio as gr
import zipfile
import tempfile
import os
import pandas as pd
from pathlib import Path
import re
from openpyxl import load_workbook
import shutil
import json
from typing import List, Dict, Tuple
import time
import traceback
import sys

# ==================== ОСНОВНЫЕ ФУНКЦИИ ====================

def extract_attributes_from_template(template_file_path: str) -> List[str]:
    """Извлекает атрибуты из файла шаблона (.xlsx)"""
    try:
        print(f"[DEBUG] Чтение шаблона: {template_file_path}")
        wb = load_workbook(template_file_path, data_only=True, read_only=True)
        ws = wb.active
        
        attributes = []
        found_start = False
        
        # Исключаемые системные атрибуты
        excluded_attributes = [
            'Наименование',
            'Наименование из системы источника',
            'Полное наименование',
            'Статус'
        ]
        
        # Ищем в первых 5 строках
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            row_values = [str(cell).strip() if cell is not None else '' for cell in row]
            row_str = ' | '.join(row_values)
            
            # Ищем начало атрибутов
            if not found_start and any('Объект данных' in str(cell) for cell in row if cell):
                print(f"[DEBUG] Начало атрибутов в строке {row_idx}: {row_str[:200]}")
                found_start = True
            
            if found_start:
                for cell in row:
                    if cell:
                        cell_str = str(cell).strip()
                        if cell_str and 'Базовая единица измерения' not in cell_str:
                            # Проверяем, не является ли атрибут исключаемым
                            is_excluded = False
                            for excluded in excluded_attributes:
                                if excluded.lower() in cell_str.lower() or cell_str.lower() in excluded.lower():
                                    is_excluded = True
                                    break
                            
                            if not is_excluded and cell_str not in attributes and cell_str:
                                attributes.append(cell_str)
                        elif 'Базовая единица измерения' in cell_str:
                            print(f"[DEBUG] Конец атрибутов: 'Базовая единица измерения'")
                            return attributes
        
        print(f"[DEBUG] Извлечено атрибутов: {len(attributes)}")
        return attributes
        
    except Exception as e:
        print(f"[ERROR] Ошибка чтения шаблона: {e}")
        return []

def process_ontology_archive(zip_file_path, progress=gr.Progress()):
    """Основная функция обработки архива"""
    if not zip_file_path:
        return None, "❌ Пожалуйста, загрузите ZIP-архив", []
    
    print(f"\n{'='*60}")
    print(f"НАЧАЛО ОБРАБОТКИ")
    print(f"Архив: {zip_file_path}")
    print(f"{'='*60}")
    
    temp_dir = tempfile.mkdtemp(prefix="ontology_")
    extracted_dir = Path(temp_dir) / "extracted"
    output_dir = Path(temp_dir) / "results"
    
    try:
        # 1. Распаковка
        progress(0, desc="📦 Распаковка архива...")
        extracted_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                zip_ref.extractall(extracted_dir)
            print(f"[SUCCESS] Архив распакован в: {extracted_dir}")
        except Exception as e:
            error_msg = f"❌ Ошибка распаковки: {str(e)}"
            print(f"[ERROR] {error_msg}")
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None, error_msg, []
        
        # 2. Анализ структуры
        progress(0.2, desc="🔍 Анализ структуры...")
        print(f"[DEBUG] Содержимое распакованной папки:")
        
        all_data = []
        missing_stats = []
        
        # Рекурсивно ищем все Excel файлы
        excel_files = list(extracted_dir.rglob("*.xlsx")) + list(extracted_dir.rglob("*.xls"))
        print(f"[INFO] Найдено Excel файлов: {len(excel_files)}")
        
        if not excel_files:
            error_msg = "❌ В архиве не найдены Excel файлы (.xlsx или .xls)"
            print(f"[ERROR] {error_msg}")
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None, error_msg, []
        
        # 3. Обработка файлов
        progress(0.3, desc="📊 Обработка данных...")
        
        processed_files = 0
        for idx, file_path in enumerate(excel_files):
            try:
                print(f"\n[PROCESSING] Файл {idx+1}/{len(excel_files)}: {file_path.name}")
                
                # Пробуем прочитать файл
                df = pd.read_excel(file_path, nrows=5)  # Читаем только первые 5 строк
                print(f"[INFO] Колонки в файле: {list(df.columns)}")
                
                # Пытаемся найти атрибуты
                attrs = extract_attributes_from_template(str(file_path))
                if attrs:
                    print(f"[SUCCESS] Найдены атрибуты: {attrs}")
                    
                    # Читаем все данные для этого файла
                    full_df = pd.read_excel(file_path)
                    
                    for attr in attrs:
                        if attr in full_df.columns:
                            values = full_df[attr].dropna().astype(str).unique()
                            for val in values:
                                if val and val.lower() not in ['nan', 'none', 'null', '']:
                                    all_data.append({
                                        'Атрибут': attr,
                                        'Файл': file_path.name,
                                        'Путь': str(file_path.relative_to(extracted_dir)),
                                        'Значение': val
                                    })
                    
                    processed_files += 1
                else:
                    print(f"[WARNING] Не удалось извлечь атрибуты из {file_path.name}")
                    
            except Exception as e:
                print(f"[ERROR] Ошибка обработки файла {file_path.name}: {e}")
                traceback.print_exc()
        
        # 4. Создание результатов
        progress(0.8, desc="💾 Сохранение результатов...")
        
        if not all_data:
            error_msg = """
            ⚠️ Не найдено данных для обработки
            
            Возможные причины:
            1. В Excel файлах нет атрибутов в строках 2-3
            2. Не найдена фраза 'Объект данных' в начале атрибутов
            3. Файлы имеют неверный формат
            
            Проверьте:
            - Что ваши файлы .xlsx/.xls содержат данные
            - Что во 2-3 строках есть названия атрибутов
            - Что есть фраза 'Объект данных' в начале списка атрибутов
            """
            print(f"[WARNING] {error_msg}")
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None, error_msg, []
        
        # Создаем выходные файлы
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Создаем сводный файл
        df_all = pd.DataFrame(all_data)
        summary_file = output_dir / "Сводные_данные.xlsx"
        df_all.to_excel(summary_file, index=False)
        
        # Создаем файлы по атрибутам
        справочники_dir = output_dir / "Вспомогательные_справочники"
        справочники_dir.mkdir(exist_ok=True)
        
        for attr in df_all['Атрибут'].unique():
            clean_name = re.sub(r'[<>:"/\\|?*]', '_', attr)
            attr_file = справочники_dir / f"{clean_name}.xlsx"
            attr_values = df_all[df_all['Атрибут'] == attr]['Значение'].unique()
            if len(attr_values) > 0:
                pd.DataFrame({'Атрибут': [attr] * len(attr_values), 'Значение': attr_values}).to_excel(attr_file, index=False)
        
        # Создаем README
        readme_file = output_dir / "README.txt"
        with open(readme_file, 'w', encoding='utf-8') as f:
            f.write("РЕЗУЛЬТАТЫ ОБРАБОТКИ\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Обработано файлов: {processed_files} из {len(excel_files)}\n")
            f.write(f"Найдено записей: {len(all_data)}\n")
            f.write(f"Уникальных атрибутов: {df_all['Атрибут'].nunique()}\n")
            f.write(f"Уникальных значений: {df_all['Значение'].nunique()}\n\n")
            f.write("Созданные файлы:\n")
            f.write("• Сводные_данные.xlsx - все извлеченные данные\n")
            f.write("• Вспомогательные_справочники/ - файлы по атрибутам\n")
            f.write("• README.txt - этот файл\n")
        
        # Создаем ZIP
        progress(0.9, desc="📦 Упаковка результатов...")
        zip_path = output_dir.parent / "результаты_обработки.zip"
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file in output_dir.rglob("*"):
                if file.is_file():
                    zipf.write(file, file.relative_to(output_dir.parent))
        
        progress(1.0, desc="✅ Готово!")
        
        # Формируем отчет
        report = f"""
{'='*50}
📊 ОТЧЕТ ОБ ОБРАБОТКЕ
{'='*50}
📅 Дата: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}

✅ УСПЕШНО ОБРАБОТАНО:
• Всего файлов в архиве: {len(excel_files)}
• Обработано файлов: {processed_files}
• Извлечено записей: {len(all_data):,}
• Уникальных атрибутов: {df_all['Атрибут'].nunique():,}
• Уникальных значений: {df_all['Значение'].nunique():,}

📁 СОЗДАННЫЕ ФАЙЛЫ:
• Сводные_данные.xlsx - все данные
• Вспомогательные_справочники/ - файлы по атрибутам
• README.txt - информация о результатах

💾 Скачайте архив со всеми файлами ниже
{'='*50}
        """
        
        print(f"\n{'='*60}")
        print(f"ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО!")
        print(f"Результаты: {zip_path}")
        print(f"{'='*60}")
        
        # ВОТ ИСПРАВЛЕНИЕ: возвращаем строку, а не кортеж
        return str(zip_path), report.strip(), [str(zip_path)]
        
    except Exception as e:
        error_msg = f"❌ Критическая ошибка: {str(e)}"
        print(f"[CRITICAL] {error_msg}")
        traceback.print_exc()
        
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except:
            pass
        
        return None, error_msg, []

# ==================== ИНТЕРФЕЙС ====================

with gr.Blocks(title="🏭 Обработчик онтологии ГРМ", theme=gr.themes.Soft()) as demo:
    
    gr.HTML("""
    <div style="text-align: center;">
        <h1 style="color: #1a237e;">🏭 Обработчик онтологии ГРМ</h1>
        <p style="color: #5c6bc0; font-size: 1.1em;">
            Извлечение атрибутов и значений из Excel файлов
        </p>
    </div>
    """)
    
    with gr.Row():
        with gr.Column(scale=2):
            gr.Markdown("""
            ### 📋 Как использовать:
            1. **Загрузите ZIP-архив** с Excel файлами (.xlsx или .xls)
            2. **Нажмите "Запустить обработку"**
            3. **Скачайте результаты**
            
            ### 🔍 Что обрабатывается:
            - Excel файлы с атрибутами в строках 2-3
            - Атрибуты начинаются с "Объект данных"
            - Исключаются системные атрибуты
            
            ### 📁 Поддерживаемая структура в архиве:
            ```
            Ваш_архив.zip/
            ├── Онтология ГРМ/                    # Корневая папка
            │   ├── Класс1/                       # Папка класса
            │   │   ├── Шаблон1/                  # Папка шаблона
            │   │   │   ├── *Шаблон.xlsx          # Файл шаблона
            │   │   │   └── *ПредЗап.xlsx         # Файл записей
            │   │   └── Шаблон2/
            │   └── Класс2/
            ```
            """)
        
        with gr.Column(scale=1):
            gr.Markdown("### 📤 Загрузка архива")
            zip_input = gr.File(
                label="ZIP архив", 
                file_types=[".zip"],
                type="filepath"
            )
            
            with gr.Row():
                process_btn = gr.Button("🚀 Запустить обработку", variant="primary", scale=2)
                clear_btn = gr.Button("🔄 Очистить", variant="secondary", scale=1)
    
    gr.Markdown("---")
    
    with gr.Row():
        with gr.Column(scale=3):
            report_output = gr.Textbox(
                label="📋 Отчет о работе",
                lines=12,
                interactive=False
            )
        
        with gr.Column(scale=1):
            gr.Markdown("### 📥 Результаты")
            result_files = gr.Files(
                label="Скачать результаты",
                interactive=False,
                height=200
            )
            
            gr.HTML("""
            <div style="background: #f5f5f5; padding: 15px; border-radius: 10px; margin-top: 20px;">
                <h4 style="margin-top: 0;">\U0001F4C8 Статистика:</h4>
                <p>После обработки здесь появится:</p>
                <ul>
                    <li>Архив с результатами</li>
                    <li>Сводный файл</li>
                    <li>Справочники по атрибутам</li>
                </ul>
            </div>
            """)
    
    def clear_all():
        return None, "", []
    
    process_btn.click(
        process_ontology_archive,
        inputs=[zip_input],
        outputs=[gr.File(visible=False), report_output, result_files],
        show_progress="full"
    )
    
    clear_btn.click(
        clear_all,
        inputs=[],
        outputs=[zip_input, report_output, result_files]
    )
    
    gr.HTML("""
    <div style="text-align: center; margin-top: 30px; padding: 15px; background: #f8f9fa; border-radius: 10px;">
        <p style="color: #666; font-size: 0.9em;">
            🏭 Обработчик онтологии ГРМ | Версия 2.0 | Gradio {version}
        </p>
    </div>
    """.format(version=gr.__version__))

# ==================== ЗАПУСК ====================

if __name__ == "__main__":
    print(f"Запуск приложения...")
    print(f"Python: {sys.version}")
    print(f"Gradio: {gr.__version__}")
    
    demo.launch(
        server_name="0.0.0.0",
        server_port=7860,
        share=False,
        debug=True
    )